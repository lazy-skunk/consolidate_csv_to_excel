import datetime
import json
import logging
import os
import sys
from logging import Logger
from logging.handlers import RotatingFileHandler
from typing import Any, Dict, List, Set

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill

_CONFIG_FILE_PATH = os.path.join("config", "config.yml")
_TARGET_FOLDERS_BASE_PATH = os.path.join("log_directory")
_LOG_FILE_PATH = os.path.join("log", "test.log")
_EXCEL_FOLDER_PATH = os.path.join("output")

_HEADER_ROW = 1
_DATA_START_ROW = _HEADER_ROW + 1
_ZERO_BASED_INDEX_OFFSET = 1
_PROCESSING_TIME_COLUMN = 3 - _ZERO_BASED_INDEX_OFFSET
_ALERT_DETAIL_COLUMN = 4 - _ZERO_BASED_INDEX_OFFSET


class CustomLogger:  # pragma: no cover
    _instance: Logger | None = None

    @classmethod
    def get_logger(
        cls,
        log_file_path: str = _LOG_FILE_PATH,
        log_level: int = logging.INFO,
        max_file_size: int = 3 * 1024 * 1024,
        backup_count: int = 2,
    ) -> Logger:
        if cls._instance is None:
            cls._instance = cls._initialize_logger(
                log_file_path, log_level, max_file_size, backup_count
            )
        return cls._instance

    @classmethod
    def _initialize_logger(
        cls,
        log_file_path: str,
        log_level: int,
        max_file_size: int,
        backup_count: int,
    ) -> Logger:
        logger = logging.getLogger(__name__)
        logger.setLevel(log_level)

        file_handler = RotatingFileHandler(
            log_file_path, maxBytes=max_file_size, backupCount=backup_count
        )
        file_handler.setLevel(log_level)
        formatter = logging.Formatter(
            fmt="%(asctime)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(log_level)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)

        return logger


class DateHandler:
    _DATE_FORMAT = "%Y%m%d"
    _DATE_DELIMITER = "~"
    _DATE_LENGTH = 8

    @classmethod
    def _parse_date(cls, input_date: str) -> datetime.datetime:
        if len(input_date) != cls._DATE_LENGTH or not input_date.isdigit():
            raise ValueError(
                f"Date must be {cls._DATE_LENGTH} digits in YYYYMMDD format."
                " For a date range, please use the format YYYYMMDD~YYYYMMDD."
                f" Input value : {input_date}"
            )

        date = datetime.datetime.strptime(input_date, DateHandler._DATE_FORMAT)

        if date > datetime.datetime.now():
            raise ValueError(
                f"Future date specified. Input value : {input_date}"
            )

        return date

    @classmethod
    def _generate_date_range(
        cls, start_date: datetime.datetime, end_date: datetime.datetime
    ) -> List[str]:
        if start_date > end_date:
            start_date, end_date = end_date, start_date

        current_date = start_date
        date_list = []
        while current_date <= end_date:
            date_list.append(current_date.strftime(cls._DATE_FORMAT))
            current_date += datetime.timedelta(days=1)

        return date_list

    @classmethod
    def get_date_range_or_yesterday(cls) -> List[str]:
        DATE_INDEX = 1

        if len(sys.argv) <= DATE_INDEX:
            yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
            return [yesterday.strftime(cls._DATE_FORMAT)]

        input_date = sys.argv[DATE_INDEX]

        if cls._DATE_DELIMITER in input_date:
            start_date_str, end_date_str = input_date.split(
                cls._DATE_DELIMITER
            )

            start_date = cls._parse_date(start_date_str)
            end_date = cls._parse_date(end_date_str)

            return cls._generate_date_range(start_date, end_date)

        date = cls._parse_date(input_date)
        return [date.strftime(cls._DATE_FORMAT)]


class ConfigLoader:
    _logger = CustomLogger.get_logger()

    def __init__(self, config_file_path: str = _CONFIG_FILE_PATH):
        self._config_file_path = config_file_path
        self._config: Dict[str, Any] = {}

    def _load_config(self) -> None:
        try:
            with open(self._config_file_path, "r") as file:
                self._config = yaml.safe_load(file)
            self._logger.info(
                f"Configuration file {self._config_file_path}"
                " loaded successfully."
            )
        except FileNotFoundError:
            raise FileNotFoundError(
                f"Configuration file {self._config_file_path} not found."
            )
        except yaml.YAMLError as e:
            raise yaml.YAMLError(
                f"Error parsing {self._config_file_path}: {e}."
            )

    def get(self, key: str, default: Any = None) -> Any:
        if not self._config:
            self._load_config()
        return self._config.get(key, default)

    def get_processing_time_threshold(self) -> int:
        threshold = self.get("processing_time_threshold_seconds")

        if isinstance(threshold, int):
            return threshold
        else:
            raise ValueError(
                "Invalid value for 'processing_time_threshold_seconds'"
                " in config file."
            )


class TargetHandler:
    @classmethod
    def get_target_prefixes(cls, config_loader: ConfigLoader) -> List[str]:
        TARGET_INDEX = 2
        if len(sys.argv) > TARGET_INDEX:
            targets = sys.argv[TARGET_INDEX].split(",")
        else:
            targets = config_loader.get("targets", [])
        return targets

    @classmethod
    def get_target_fullnames(cls, target_prefixes: List[str]) -> List[str]:
        target_fullnames = []
        target_folders = os.listdir(_TARGET_FOLDERS_BASE_PATH)

        for target_prefix in target_prefixes:
            matched_target_fullnames = [
                target_folder
                for target_folder in target_folders
                if target_folder.startswith(target_prefix)
            ]

            if matched_target_fullnames:
                target_fullnames.extend(matched_target_fullnames)
            else:
                raise ValueError(
                    f"No folder starting with target prefix '{target_prefix}'"
                    " was found in the log directory."
                )

        return target_fullnames


class CSVPathMapper:
    @staticmethod
    def get_targets_and_csv_path_by_dates(
        date_range: List[str], target_fullnames: List[str]
    ) -> Dict[str, Dict[str, str | None]]:
        targets_and_csv_path_by_dates = {}

        for date in date_range:
            csv_paths = {
                target_fullname: FileUtility.get_csv_path(
                    os.path.join(_TARGET_FOLDERS_BASE_PATH, target_fullname),
                    date,
                )
                for target_fullname in target_fullnames
            }

            targets_and_csv_path_by_dates[date] = csv_paths

        return targets_and_csv_path_by_dates


class FileUtility:
    @staticmethod
    def create_excel_path(date: str, suffix: str) -> str:
        excel_name = f"{date}_{suffix}.xlsx"
        excel_path = os.path.join(_EXCEL_FOLDER_PATH, date, excel_name)
        return excel_path

    @staticmethod
    def create_directory(file_path: str) -> None:
        directory_for_file = os.path.dirname(file_path)
        os.makedirs(directory_for_file, exist_ok=True)

    @staticmethod
    def get_csv_path(target_folder_path: str, date: str) -> str | None:
        csv_name = f"test_{date}.csv"
        csv_path = os.path.join(target_folder_path, csv_name)

        if os.path.exists(csv_path):
            return csv_path
        else:
            return None


class CSVConsolidator:
    _logger = CustomLogger.get_logger()

    def __init__(self, writer: pd.ExcelWriter, workbook: Workbook) -> None:
        self._writer = writer
        self._workbook = workbook

        self._merge_failed_hosts: set[str] = set()

    def _create_sentinel_sheet(self) -> None:
        pd.DataFrame({"A": ["SENTINEL_SHEET"]}).to_excel(
            self._writer,
            sheet_name="SENTINEL_SHEET",
            index=False,
            header=False,
        )

    def _create_sheet_from_csv(self, target_name: str, csv_path: str) -> None:
        try:
            df = pd.read_csv(csv_path)
            df.to_excel(self._writer, sheet_name=target_name, index=False)
        except Exception as e:
            self._logger.error(f"Failed to read CSV file at {csv_path}: {e}")
            self._merge_failed_hosts.add(target_name)

    def _create_no_csv_sheet(self, target_name: str) -> None:
        df_for_no_csv = pd.DataFrame({"A": ["No CSV file found."]})
        df_for_no_csv.to_excel(
            self._writer, sheet_name=target_name, index=False, header=False
        )

        _TRANSPARENT = "FF"
        _GRAY = "7F7F7F"
        _GRAY_WITH_TRANSPARENT = _TRANSPARENT + _GRAY
        self._writer.sheets[target_name].sheet_properties.tabColor = (
            _GRAY_WITH_TRANSPARENT
        )

    def _create_sheets(
        self, filtered_targets_and_csv_path: dict[str, str | None]
    ) -> None:
        total_targets = len(filtered_targets_and_csv_path)

        for current_target_number, (target_name, csv_path) in enumerate(
            filtered_targets_and_csv_path.items(), start=1
        ):
            if csv_path:
                self._create_sheet_from_csv(target_name, csv_path)
            else:
                self._create_no_csv_sheet(target_name)

            self._logger.info(
                f"Added sheet: {target_name}."
                f" ({current_target_number}/{total_targets})"
            )

    def _delete_sentinel_sheet(self) -> None:
        if "SENTINEL_SHEET" in self._workbook.sheetnames:
            del self._workbook["SENTINEL_SHEET"]

    def consolidate_csvs_to_excel(
        self, filtered_targets_and_csv_path: dict[str, str | None]
    ) -> None:
        self._logger.info("Starting to merge.")

        self._create_sentinel_sheet()
        self._create_sheets(filtered_targets_and_csv_path)
        self._delete_sentinel_sheet()

        self._logger.info("Merging completed.")

    def get_merge_failed_hosts(self) -> Dict[str, set[str]]:
        return {"merge_failed_hosts": self._merge_failed_hosts}


class ExcelAnalyzer:
    _logger = CustomLogger.get_logger()
    _TRANSPARENT = "FF"
    _YELLOW = "FFFF7F"
    _GRAY = "7F7F7F"
    _YELLOW_WITH_TRANSPARENT = _TRANSPARENT + _YELLOW
    _GRAY_WITH_TRANSPARENT = _TRANSPARENT + _GRAY

    def __init__(self, workbook: Workbook) -> None:
        self._workbook = workbook
        self._hosts_with_threshold_exceedance: set[str] = set()
        self._hosts_with_anomaly_value: set[str] = set()

    @staticmethod
    def _highlight_cell(cell: Cell, color_code: str) -> None:
        pattern_fill = PatternFill(start_color=color_code, fill_type="solid")
        cell.fill = pattern_fill

    @staticmethod
    def _calculate_color_based_on_excess_ratio(
        processing_time_seconds: int, threshold: int
    ) -> str:
        excess_ratio = (processing_time_seconds - threshold) / threshold
        clamped_excess_ratio = min(excess_ratio, 1)

        MAX_GREEN_VALUE = 255
        MIN_GREEN_VALUE = MAX_GREEN_VALUE / 2
        green_value = int(
            MAX_GREEN_VALUE
            - (MAX_GREEN_VALUE - MIN_GREEN_VALUE) * clamped_excess_ratio
        )

        green_hex_value = f"{green_value:02X}"
        color_code = f"FF{green_hex_value}7F"

        return color_code

    def _check_and_highlight_processing_time(
        self,
        processing_time_cell: Cell,
        threshold: int,
    ) -> bool:
        processing_time_value = processing_time_cell.value

        if processing_time_value:
            try:
                processing_time_seconds = int(
                    processing_time_value.rstrip("s")
                )

                if processing_time_seconds >= threshold:
                    color_code = self._calculate_color_based_on_excess_ratio(
                        processing_time_seconds, threshold
                    )
                    self._highlight_cell(processing_time_cell, color_code)
                    return True
            except ValueError:
                self._logger.warning(
                    f"Invalid processing time value: {processing_time_value}"
                )

        return False

    def _check_and_highlight_alert_detail(
        self,
        alert_detail_cell: Cell,
    ) -> bool:
        alert_detail_value = alert_detail_cell.value

        if alert_detail_value:
            try:
                alert_detail_data = json.loads(alert_detail_value)
                if any(
                    item.get("random_key") is True
                    for item in alert_detail_data
                ):
                    self._highlight_cell(alert_detail_cell, self._YELLOW)
                    return True
            except json.JSONDecodeError:
                self._logger.warning(
                    f"Invalid JSON format found: {alert_detail_value}"
                )
        return False

    def _log_detected_anomalies(self, host_name: str) -> None:
        if host_name in self._hosts_with_threshold_exceedance:
            self._logger.warning(
                "Exceeded processing time threshold detected"
                f" for host: {host_name}."
            )

        if host_name in self._hosts_with_anomaly_value:
            self._logger.warning(
                f"Anomaly value detected for host: {host_name}."
            )

    def highlight_cells_and_sheet_tab_by_criteria(
        self, threshold: int
    ) -> None:
        self._logger.info("Starting to highlight.")
        total_sheets = len(self._workbook.sheetnames)

        for current_sheet_number, host_name in enumerate(
            self._workbook.sheetnames, start=1
        ):
            sheet = self._workbook[host_name]
            has_highlighted_cell = False

            for row in sheet.iter_rows(min_row=_DATA_START_ROW):
                processing_time_cell = row[_PROCESSING_TIME_COLUMN]
                alert_detail_cell = row[_ALERT_DETAIL_COLUMN]

                if self._check_and_highlight_processing_time(
                    processing_time_cell, threshold
                ):
                    self._hosts_with_threshold_exceedance.add(host_name)
                    has_highlighted_cell = True

                if self._check_and_highlight_alert_detail(alert_detail_cell):
                    self._hosts_with_anomaly_value.add(host_name)
                    has_highlighted_cell = True

            if has_highlighted_cell:
                sheet.sheet_properties.tabColor = self._YELLOW_WITH_TRANSPARENT
                self._log_detected_anomalies(host_name)

            self._logger.info(
                f"Analyzed sheet: {host_name}."
                f" ({current_sheet_number}/{total_sheets})"
            )
        self._logger.info("Highlighting completed.")

    def _create_new_order(self) -> List[str]:
        yellow_sheets = []
        gray_sheets = []
        other_sheets = []

        for sheet_name in self._workbook.sheetnames:
            sheet_tab_color = self._workbook[
                sheet_name
            ].sheet_properties.tabColor

            if sheet_tab_color is None:
                other_sheets.append(sheet_name)
            else:
                sheet_color_value = sheet_tab_color.value

                if sheet_color_value == self._YELLOW_WITH_TRANSPARENT:
                    yellow_sheets.append(sheet_name)
                elif sheet_color_value == self._GRAY_WITH_TRANSPARENT:
                    gray_sheets.append(sheet_name)

        return yellow_sheets + other_sheets + gray_sheets

    def reorder_sheets_by_color(self) -> None:
        self._logger.info("Starting to reorder.")
        new_order = self._create_new_order()

        total_sheets = len(self._workbook.sheetnames)
        for current_sheet_number, sheet_name in enumerate(new_order, start=1):
            self._workbook.move_sheet(sheet_name, total_sheets)
            self._logger.info(
                f"Reordered sheet: {sheet_name}."
                f" ({current_sheet_number}/{total_sheets})"
            )
        self._logger.info("Reordering completed.")

    def get_analysis_results(self) -> Dict[str, set[str]]:
        return {
            "hosts_with_threshold_exceedance": self._hosts_with_threshold_exceedance,  # noqa E501
            "hosts_with_anomaly_value": self._hosts_with_anomaly_value,
        }


class ProcessingSummary:
    _logger = CustomLogger.get_logger()

    def __init__(self) -> None:
        self.daily_summaries: Dict[str, List[str]] = {}
        self.daily_processing_results: Dict[str, Dict[str, Set[str]]] = {}

    def add_missing_csv_info(
        self,
        targets_and_csv_path_by_dates: Dict[str, Dict[str, str | None]],
    ) -> None:
        for (
            date,
            targets_and_csv_path,
        ) in targets_and_csv_path_by_dates.items():
            if all(
                csv_path is None for csv_path in targets_and_csv_path.values()
            ):
                self.daily_summaries.setdefault(date, []).append(
                    "No CSV files found."
                )
            else:
                missing_targets = [
                    target_fullname
                    for target_fullname, csv_path in targets_and_csv_path.items()  # noqa E501
                    if csv_path is None
                ]
                if missing_targets:
                    self.daily_summaries.setdefault(date, []).append(
                        f"Partial data loss, missing hosts: {missing_targets}"
                    )

    def save_daily_processing_results(
        self,
        date: str,
        csv_consolidator: CSVConsolidator,
        excel_analyzer: ExcelAnalyzer,
    ) -> None:
        merge_failed_hosts = csv_consolidator.get_merge_failed_hosts()
        analysis_results = excel_analyzer.get_analysis_results()

        self.daily_processing_results.setdefault(
            date,
            {
                "merge_failed_hosts": set(),
                "hosts_with_threshold_exceedance": set(),
                "hosts_with_anomaly_value": set(),
            },
        )

        self.daily_processing_results[date]["merge_failed_hosts"].update(
            merge_failed_hosts["merge_failed_hosts"]
        )

        self.daily_processing_results[date][
            "hosts_with_threshold_exceedance"
        ].update(analysis_results["hosts_with_threshold_exceedance"])

        self.daily_processing_results[date]["hosts_with_anomaly_value"].update(
            analysis_results["hosts_with_anomaly_value"]
        )

    def _summarize_daily_processing_results(self) -> None:
        for date, summary in self.daily_processing_results.items():
            day_summary = []

            if summary.get("hosts_with_threshold_exceedance"):
                hosts_with_threshold_exceedance = summary[
                    "hosts_with_threshold_exceedance"
                ]
                day_summary.append(
                    "Exceeded threshold detected for hosts:"
                    f" {hosts_with_threshold_exceedance}"
                )

            if summary.get("hosts_with_anomaly_value"):
                hosts_with_anomaly_value = summary["hosts_with_anomaly_value"]
                day_summary.append(
                    "Anomaly value detected for hosts:"
                    f" {hosts_with_anomaly_value}"
                )

            if summary.get("merge_failed_hosts"):
                merge_failed_hosts = summary["merge_failed_hosts"]
                day_summary.append(f"Merge failed hosts: {merge_failed_hosts}")

            self.daily_summaries.setdefault(date, []).extend(day_summary)

    def log_daily_summaries(self) -> None:
        self._logger.info("Starting to log summary.")
        self._summarize_daily_processing_results()

        for date in sorted(self.daily_summaries.keys()):
            self._logger.info(f"Summary for {date}:")

            if not self.daily_summaries[date]:
                self._logger.info("No anomalies detected.")
            else:
                for summary_item in self.daily_summaries[date]:
                    self._logger.warning(f"{summary_item}")

        self._logger.info("Finished logging summary.")


def main() -> None:
    try:
        logger = CustomLogger.get_logger()
        logger.info("Process started.")

        date_range = DateHandler.get_date_range_or_yesterday()
        config_loader = ConfigLoader(_CONFIG_FILE_PATH)
        target_prefixes = TargetHandler.get_target_prefixes(config_loader)
        target_fullnames = TargetHandler.get_target_fullnames(target_prefixes)
        targets_and_csv_path_by_dates = (
            CSVPathMapper.get_targets_and_csv_path_by_dates(
                date_range, target_fullnames
            )
        )
        processing_time_threshold = (
            config_loader.get_processing_time_threshold()
        )
        processing_summary = ProcessingSummary()
        processing_summary.add_missing_csv_info(targets_and_csv_path_by_dates)

        for (
            date,
            targets_and_csv_path,
        ) in targets_and_csv_path_by_dates.items():
            if all(
                csv_path is None for csv_path in targets_and_csv_path.values()
            ):
                logger.warning(f"No CSV files found for date {date}.")
                continue

            for target_prefix in target_prefixes:
                extracted_targets_and_csv_path = {
                    target_fullname: csv_path
                    for target_fullname, csv_path in targets_and_csv_path.items()  # noqa E501
                    if target_fullname.startswith(target_prefix)
                }

                if all(
                    csv_path is None
                    for csv_path in extracted_targets_and_csv_path.values()
                ):
                    logger.warning(
                        "No CSV files found for"
                        f" target prefix '{target_prefix}' on date {date}."
                    )
                    continue

                excel_path = FileUtility.create_excel_path(date, target_prefix)
                FileUtility.create_directory(excel_path)

                logger.info(f"Starting to create '{excel_path}'.")
                with pd.ExcelWriter(
                    excel_path, engine="openpyxl", mode="w"
                ) as writer:
                    workbook = writer.book

                    csv_consolidator = CSVConsolidator(writer, workbook)
                    csv_consolidator.consolidate_csvs_to_excel(
                        extracted_targets_and_csv_path
                    )

                    excel_analyzer = ExcelAnalyzer(workbook)
                    excel_analyzer.highlight_cells_and_sheet_tab_by_criteria(
                        processing_time_threshold
                    )
                    excel_analyzer.reorder_sheets_by_color()

                    logger.info(f"Saving '{excel_path}'.")

                processing_summary.save_daily_processing_results(
                    date,
                    csv_consolidator,
                    excel_analyzer,
                )
                logger.info(f"Finished creating '{excel_path}'.")

        processing_summary.log_daily_summaries()
        logger.info("Process completed.")
    except Exception as e:
        logger.error(f"An error occured: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
