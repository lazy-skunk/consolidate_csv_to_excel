import datetime
import json
import logging
import os
import sys
from logging import Logger
from logging.handlers import RotatingFileHandler
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill

_CONFIG_FILE_PATH = os.path.join("config", "config.yml")
_TARGET_FOLDERS_BASE_PATH = os.path.join("log_directory")
_LOG_FILE_PATH = os.path.join("log", "test.log")
_EXCEL_FOLDER_PATH = os.path.join("output", "excel")


class CustomLogger:  # pragma: no cover
    def __init__(
        self,
        log_file_path: str = _LOG_FILE_PATH,
        log_level: int = logging.INFO,
        max_file_size: int = 3 * 1024 * 1024,
        backup_count: int = 2,
    ) -> None:
        self._logger = logging.getLogger(__name__)
        self._logger.setLevel(log_level)

        file_handler = RotatingFileHandler(
            log_file_path, maxBytes=max_file_size, backupCount=backup_count
        )
        file_handler.setLevel(log_level)
        formatter = logging.Formatter(
            fmt="%(asctime)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        file_handler.setFormatter(formatter)
        self._logger.addHandler(file_handler)

        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(log_level)
        console_handler.setFormatter(formatter)
        self._logger.addHandler(console_handler)

    @property
    def get_logger(self) -> Logger:
        return self._logger


class DateHandler:
    _DATE_FORMAT = "%Y%m%d"
    _DATE_DELIMITER = "~"
    _DATE_LENGTH = 8

    def __init__(self, logger: Logger) -> None:
        self._logger = logger

    def _parse_date(self, input_date: str) -> Optional[datetime.datetime]:
        if len(input_date) != self._DATE_LENGTH or not input_date.isdigit():
            error_message = (
                f"Date must be {self._DATE_LENGTH} digits in YYYYMMDD format :"
                f" {input_date}."
                " For a date range, please use the format YYYYMMDD~YYYYMMDD."
            )
            self._logger.error(error_message)
            raise ValueError(error_message)

        date = datetime.datetime.strptime(input_date, DateHandler._DATE_FORMAT)
        if date > datetime.datetime.now():
            error_message = f"Future date specified: {input_date}."
            self._logger.error(error_message)
            raise ValueError(error_message)

        return date

    def _generate_date_range(
        self, start_date: datetime.datetime, end_date: datetime.datetime
    ) -> List[str]:
        if start_date > end_date:
            start_date, end_date = end_date, start_date

        current_date = start_date
        date_list = []
        while current_date <= end_date:
            date_list.append(current_date.strftime(DateHandler._DATE_FORMAT))
            current_date += datetime.timedelta(days=1)

        return date_list

    def get_date_range_or_yesterday(self) -> List[str]:
        DATE_INDEX = 1

        if len(sys.argv) > DATE_INDEX:
            input_date = sys.argv[DATE_INDEX]

            if self._DATE_DELIMITER in input_date:
                dates = input_date.split(self._DATE_DELIMITER)

                if len(dates) != 2:
                    error_message = (
                        f"Invalid date range format specified: {input_date}. "
                        "Please use the format YYYYMMDD~YYYYMMDD."
                    )
                    self._logger.error(error_message)
                    raise ValueError(error_message)

                start_date_str, end_date_str = dates
                start_date = self._parse_date(start_date_str)
                end_date = self._parse_date(end_date_str)

                if start_date and end_date:
                    return self._generate_date_range(start_date, end_date)
            else:
                date = self._parse_date(input_date)
                if date:
                    return [date.strftime(DateHandler._DATE_FORMAT)]

        yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
        return [yesterday.strftime(DateHandler._DATE_FORMAT)]


class ConfigLoader:
    def __init__(
        self, logger: Logger, config_file_path: str = _CONFIG_FILE_PATH
    ) -> None:
        self._logger = logger
        self._config_file_path = config_file_path
        self._config = self._load_config()

    def _load_config(self) -> Dict[str, Any]:
        try:
            with open(self._config_file_path, "r") as file:
                config = yaml.safe_load(file)
            self._logger.info(
                f"Configuration file {self._config_file_path}"
                " loaded successfully."
            )
            return config
        except FileNotFoundError:
            error_message = (
                f"Configuration file {self._config_file_path} not found."
            )
            self._logger.error(error_message)
            raise
        except yaml.YAMLError as e:
            error_message = f"Error parsing {self._config_file_path} : {e}."
            self._logger.error(error_message)
            raise

    def get(self, key: str, default: Any = None) -> Any:
        return self._config.get(key, default)

    def get_processing_time_threshold(self) -> int:
        threshold = self.get("processing_time_threshold_seconds")

        if isinstance(threshold, int):
            return threshold
        else:
            error_message = (
                "Invalid value for 'processing_time_threshold_seconds'"
                " in config file. Please provide a valid integer value."
            )
            self._logger.error(error_message)
            raise ValueError(error_message)


class TargetHandler:
    def __init__(self, config_loader: ConfigLoader, logger: Logger) -> None:
        self.config_loader = config_loader
        self._logger = logger

    def get_targets(self) -> List[str]:
        TARGET = 2
        if len(sys.argv) > 2:
            targets = sys.argv[TARGET].split(",")
        else:
            targets = self.config_loader.get("targets", [])

        return targets

    def get_existing_target_fullnames(self, targets: List[str]) -> List[str]:
        host_fullnames = []
        host_folders = os.listdir(_TARGET_FOLDERS_BASE_PATH)

        for target in targets:
            matched_host_names = [
                host_folder
                for host_folder in host_folders
                if host_folder.startswith(target)
            ]

            if matched_host_names:
                host_fullnames.extend(matched_host_names)
            else:
                self._logger.warning(
                    f"No folder starting with target '{target}'"
                    " was found in the log directory."
                )

        if not host_fullnames:
            error_message = "No valid targets found."
            self._logger.error(error_message)
            raise ValueError(error_message)

        return host_fullnames


class FileUtility:
    @staticmethod
    def create_file_name_suffix(targets: List[str]) -> str:
        if len(sys.argv) > 2:
            return "_".join(targets)
        else:
            return "config"

    @staticmethod
    def create_excel_path(date: str, suffix: str) -> str:
        excel_name = f"{date}_{suffix}.xlsx"
        excel_path = os.path.join(_EXCEL_FOLDER_PATH, date, excel_name)
        return excel_path

    @staticmethod
    def create_directory(file_path: str) -> None:
        directory_for_file = os.path.dirname(file_path)
        os.makedirs(directory_for_file, exist_ok=True)

    @staticmethod
    def get_merged_csv_path(
        target_folder_path: str, date: str
    ) -> Optional[str]:
        csv_name = f"test_{date}.csv"
        csv_path = os.path.join(target_folder_path, csv_name)

        if os.path.exists(csv_path):
            return csv_path
        else:
            return None


class CSVConsolidator:
    def __init__(
        self, writer: pd.ExcelWriter, workbook: Workbook, logger: Logger
    ) -> None:
        self._writer = writer
        self._workbook = workbook
        self._logger = logger

        self._copied_count = 0
        self._no_csv_count = 0
        self._failed_count = 0
        self._failed_hosts: List[str] = []

    def create_sentinel_sheet(self) -> None:
        pd.DataFrame({"A": ["SENTINEL_SHEET"]}).to_excel(
            self._writer,
            sheet_name="SENTINEL_SHEET",
            index=False,
            header=False,
        )

    def _copy_csv_to_excel(self, csv_path: str, target_name: str) -> None:
        try:
            df = pd.read_csv(csv_path)
            df.to_excel(self._writer, sheet_name=target_name, index=False)
            self._copied_count += 1
        except Exception as e:
            self._logger.error(f"Failed to read CSV file at {csv_path}: {e}")
            self._failed_count += 1
            self._failed_hosts.append(target_name)

    def _create_no_csv_sheet_to_excel(self, target_name: str) -> None:
        df_for_not_found = pd.DataFrame({"A": ["No CSV file found."]})
        df_for_not_found.to_excel(
            self._writer, sheet_name=target_name, index=False, header=False
        )

        GRAY = "7F7F7F"
        self._writer.sheets[target_name].sheet_properties.tabColor = GRAY

        self._no_csv_count += 1

    def _add_sheet_for_target(
        self, target_folder_path: str, date: str
    ) -> None:
        csv_path = FileUtility.get_merged_csv_path(target_folder_path, date)
        target_name = os.path.basename(target_folder_path)

        if csv_path:
            self._copy_csv_to_excel(csv_path, target_name)
        else:
            self._create_no_csv_sheet_to_excel(target_name)

    def search_and_append_csv_to_excel(
        self, date: str, target_fullnames: List[str]
    ) -> None:
        total_targets = len(target_fullnames)
        for current_target_number, target_name in enumerate(
            target_fullnames, start=1
        ):
            target_folder_path = os.path.join(
                _TARGET_FOLDERS_BASE_PATH, target_name
            )
            self._add_sheet_for_target(target_folder_path, date)
            self._logger.info(
                f"Added sheet: {target_name}."
                f" ({current_target_number}/{total_targets})"
            )

    def delete_sentinel_sheet(self) -> None:
        if "SENTINEL_SHEET" in self._workbook.sheetnames:
            del self._workbook["SENTINEL_SHEET"]
            return

    def get_summary(self) -> Dict[str, int | List[str]]:
        return {
            "copied": self._copied_count,
            "no_csv": self._no_csv_count,
            "failed": self._failed_count,
            "failed_hosts": self._failed_hosts,
        }


class ExcelAnalyzer:
    _TRANSPARENT = "00"
    _YELLOW = "FFFF7F"
    _GRAY = "7F7F7F"
    _YELLOW_WITH_TRANSPARENT = _TRANSPARENT + _YELLOW
    _GRAY_WITH_TRANSPARENT = _TRANSPARENT + _GRAY

    def __init__(self, workbook: Workbook, logger: Logger) -> None:
        self._workbook = workbook
        self._logger = logger
        self._hosts_to_check: set[str] = set()

    def _highlight_cell(self, cell: Cell, color_code: str) -> None:
        pattern_fill = PatternFill(start_color=color_code, fill_type="solid")
        cell.fill = pattern_fill

    def _calculate_color_based_on_excess_ratio(
        self, processing_time_seconds: int, threshold: int
    ) -> str:
        excess_ratio = (processing_time_seconds - threshold) / threshold
        clamped_excess_ratio = min(excess_ratio, 1)

        MAX_GREEN_VALUE = 255
        MIN_GREEN_VALUE = MAX_GREEN_VALUE / 2
        green_value = int(
            MAX_GREEN_VALUE
            - (MAX_GREEN_VALUE - MIN_GREEN_VALUE) * clamped_excess_ratio
        )

        green_hex_value = f"{green_value:02X}"
        color_code = f"FF{green_hex_value}7F"

        return color_code

    def _check_and_highlight_processing_time(
        self,
        row: Tuple[Cell, ...],
        processing_time_column: int,
        threshold: int,
    ) -> bool:
        processing_time_cell = row[processing_time_column]
        processing_time_value = processing_time_cell.value

        if processing_time_value:
            try:
                processing_time_seconds = int(
                    processing_time_value.rstrip("s")
                )

                if processing_time_seconds >= threshold:
                    color_code = self._calculate_color_based_on_excess_ratio(
                        processing_time_seconds, threshold
                    )
                    self._highlight_cell(processing_time_cell, color_code)
                    return True
            except ValueError:
                self._logger.warning(
                    f"Invalid processing time value: {processing_time_value}"
                )

        return False

    def _check_and_highlight_alert_detail(
        self,
        row: tuple[Cell, ...],
        alert_detail_column: int,
    ) -> bool:
        alert_detail_cell = row[alert_detail_column]
        alert_detail_value = alert_detail_cell.value

        if alert_detail_value:
            try:
                alert_detail_data = json.loads(alert_detail_value)
                if any(
                    item.get("random_key") is True
                    for item in alert_detail_data
                ):
                    self._highlight_cell(alert_detail_cell, self._YELLOW)
                    return True
            except json.JSONDecodeError:
                self._logger.warning(
                    f"Invalid JSON format found: {alert_detail_value}"
                )
        return False

    def highlight_cells_and_sheet_tab_by_criteria(
        self, threshold: int
    ) -> None:
        HEADER_ROW = 1
        DATA_START_ROW = HEADER_ROW + 1
        ZERO_BASED_INDEX_OFFSET = 1
        PROCESSING_TIME_COLUMN = 3 - ZERO_BASED_INDEX_OFFSET
        ALERT_DETAIL_COLUMN = 4 - ZERO_BASED_INDEX_OFFSET
        total_sheets = len(self._workbook.sheetnames)

        for current_sheet_number, host_name in enumerate(
            self._workbook.sheetnames, start=1
        ):
            sheet = self._workbook[host_name]
            has_highlighted_cell = False

            for row in sheet.iter_rows(min_row=DATA_START_ROW):
                processing_time_highlighted = (
                    self._check_and_highlight_processing_time(
                        row, PROCESSING_TIME_COLUMN, threshold
                    )
                )
                alert_detail_highlighted = (
                    self._check_and_highlight_alert_detail(
                        row, ALERT_DETAIL_COLUMN
                    )
                )

                if processing_time_highlighted or alert_detail_highlighted:
                    has_highlighted_cell = True

            if has_highlighted_cell:
                self._hosts_to_check.add(host_name)
                sheet.sheet_properties.tabColor = self._YELLOW

            self._logger.info(
                f"Analyzed sheet: {host_name}."
                f" ({current_sheet_number}/{total_sheets})"
            )

    def _create_new_order(self) -> List[str]:
        yellow_sheets = []
        gray_sheets = []
        other_sheets = []

        for sheet_name in self._workbook.sheetnames:
            sheet_tab_color = self._workbook[
                sheet_name
            ].sheet_properties.tabColor
            if sheet_tab_color is None:
                other_sheets.append(sheet_name)
            else:
                sheet_color_value = sheet_tab_color.value
                if sheet_color_value == self._YELLOW_WITH_TRANSPARENT:
                    yellow_sheets.append(sheet_name)
                elif sheet_color_value == self._GRAY_WITH_TRANSPARENT:
                    gray_sheets.append(sheet_name)
        return yellow_sheets + other_sheets + gray_sheets

    def reorder_sheets_by_color(self) -> None:
        new_order = self._create_new_order()

        total_sheets = len(self._workbook.sheetnames)
        for current_sheet_number, sheet_name in enumerate(new_order, start=1):
            self._workbook.move_sheet(sheet_name, total_sheets)
            self._logger.info(
                f"Reordered sheet: {sheet_name}."
                f" ({current_sheet_number}/{total_sheets})"
            )

    def get_hosts_to_check(self) -> set:
        return self._hosts_to_check


def _save_daily_summary(  # pragma: no cover
    daily_summaries: Dict[str, Dict[str, int | List[str] | Set[str]]],
    date: str,
    consolidator: CSVConsolidator,
    excel_analyzer: ExcelAnalyzer,
) -> None:
    consolidator_summary = consolidator.get_summary()
    hosts_to_check = excel_analyzer.get_hosts_to_check()

    daily_summaries[date] = {
        **consolidator_summary,
        "hosts_to_check": hosts_to_check,
    }


def _log_daily_summaries(  # pragma: no cover
    logger: Logger,
    daily_summaries: Dict[str, Dict[str, int | List[str] | Set[str]]],
) -> None:
    logger.info("Logging daily summaries for each date:")
    for date, summary in daily_summaries.items():
        logger.info(
            f"Date: {date} - Copied: {summary['copied']},"
            f" No CSV: {summary['no_csv']}. "
        )

        if summary.get("hosts_to_check"):
            hosts_to_check = summary["hosts_to_check"]
            if isinstance(hosts_to_check, set):
                logger.warning(
                    f"Date: {date} - hosts to check:"
                    f" {', '.join(hosts_to_check)}"
                )
            else:
                logger.warning(
                    f"Date: {date} - hosts to check: {hosts_to_check}"
                )

        if summary.get("failed") and summary.get("failed_hosts"):
            failed_hosts = summary["failed_hosts"]
            if isinstance(failed_hosts, list):
                logger.error(
                    f"Date: {date} - Failed: {summary['failed']},"
                    f" Failed Hosts: {', '.join(failed_hosts)}"
                )
            else:
                logger.error(
                    f"Date: {date} - Failed: {summary['failed']},"
                    f" Failed Hosts: {failed_hosts}"
                )


if __name__ == "__main__":  # pragma: no cover
    try:
        logger = CustomLogger(_LOG_FILE_PATH).get_logger
        logger.info("Process started.")

        date_handler = DateHandler(logger)
        config_loader = ConfigLoader(logger, _CONFIG_FILE_PATH)
        target_handler = TargetHandler(config_loader, logger)

        date_range = date_handler.get_date_range_or_yesterday()
        targets = target_handler.get_targets()
        target_fullnames = target_handler.get_existing_target_fullnames(
            targets
        )
        processing_time_threshold = (
            config_loader.get_processing_time_threshold()
        )
        daily_summaries: Dict[str, Dict[str, int | List[str] | Set[str]]] = {}

        file_name_suffix = FileUtility.create_file_name_suffix(targets)
        for date in date_range:
            excel_path = FileUtility.create_excel_path(date, file_name_suffix)
            FileUtility.create_directory(excel_path)

            with pd.ExcelWriter(
                excel_path, engine="openpyxl", mode="w"
            ) as writer:
                workbook = writer.book

                consolidator = CSVConsolidator(writer, workbook, logger)
                consolidator.create_sentinel_sheet()
                consolidator.search_and_append_csv_to_excel(
                    date, target_fullnames
                )
                consolidator.delete_sentinel_sheet()

                excel_analyzer = ExcelAnalyzer(workbook, logger)
                excel_analyzer.highlight_cells_and_sheet_tab_by_criteria(
                    processing_time_threshold
                )
                excel_analyzer.reorder_sheets_by_color()

            _save_daily_summary(
                daily_summaries, date, consolidator, excel_analyzer
            )

        _log_daily_summaries(logger, daily_summaries)
        logger.info("Process completed.")
    except Exception as e:
        logger.error(f"An error occured: {e}", exc_info=True)
        sys.exit(1)
