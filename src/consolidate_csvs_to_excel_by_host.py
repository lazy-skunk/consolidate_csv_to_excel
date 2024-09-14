import datetime
import json
import logging
import os
import sys
from logging import Logger
from logging.handlers import RotatingFileHandler
from typing import Any, Dict, List, Set

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill

_CONFIG_FILE_PATH = os.path.join("config", "config.yml")
_TARGET_FOLDERS_BASE_PATH = os.path.join("log_directory")
_LOG_FILE_PATH = os.path.join("log", "test.log")
_EXCEL_FOLDER_PATH = os.path.join("output")

_HEADER_ROW = 1
_DATA_START_ROW = _HEADER_ROW + 1
_ZERO_BASED_INDEX_OFFSET = 1
_PROCESSING_TIME_COLUMN = 3 - _ZERO_BASED_INDEX_OFFSET
_ALERT_DETAIL_COLUMN = 4 - _ZERO_BASED_INDEX_OFFSET


class CustomLogger:  # pragma: no cover
    _instance: Logger | None = None

    @classmethod
    def get_logger(
        cls,
        log_file_path: str = _LOG_FILE_PATH,
        log_level: int = logging.INFO,
        max_file_size: int = 3 * 1024 * 1024,
        backup_count: int = 2,
    ) -> Logger:
        if cls._instance is None:
            cls._instance = cls._initialize_logger(
                log_file_path, log_level, max_file_size, backup_count
            )
        return cls._instance

    @classmethod
    def _initialize_logger(
        cls,
        log_file_path: str,
        log_level: int,
        max_file_size: int,
        backup_count: int,
    ) -> Logger:
        logger = logging.getLogger(__name__)
        logger.setLevel(log_level)

        file_handler = RotatingFileHandler(
            log_file_path, maxBytes=max_file_size, backupCount=backup_count
        )
        file_handler.setLevel(log_level)
        formatter = logging.Formatter(
            fmt="%(asctime)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(log_level)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)

        return logger


class DateHandler:
    _DATE_FORMAT = "%Y%m%d"
    _DATE_DELIMITER = "~"
    _DATE_LENGTH = 8

    @classmethod
    def _parse_date(cls, input_date: str) -> datetime.datetime:
        if len(input_date) != cls._DATE_LENGTH or not input_date.isdigit():
            raise ValueError(
                f"Date must be {cls._DATE_LENGTH} digits in YYYYMMDD format."
                " For a date range, please use the format YYYYMMDD~YYYYMMDD."
                f" Input value : {input_date}"
            )

        date = datetime.datetime.strptime(input_date, DateHandler._DATE_FORMAT)

        if date > datetime.datetime.now():
            raise ValueError(
                f"Future date specified. Input value : {input_date}"
            )

        return date

    @classmethod
    def _generate_date_range(
        cls, start_date: datetime.datetime, end_date: datetime.datetime
    ) -> List[str]:
        if start_date > end_date:
            start_date, end_date = end_date, start_date

        current_date = start_date
        date_list = []
        while current_date <= end_date:
            date_list.append(current_date.strftime(cls._DATE_FORMAT))
            current_date += datetime.timedelta(days=1)

        return date_list

    @classmethod
    def get_date_range_or_yesterday(cls) -> List[str]:
        DATE_INDEX = 1

        if len(sys.argv) <= DATE_INDEX:
            yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
            return [yesterday.strftime(cls._DATE_FORMAT)]

        input_date = sys.argv[DATE_INDEX]

        if cls._DATE_DELIMITER in input_date:
            start_date_str, end_date_str = input_date.split(
                cls._DATE_DELIMITER
            )

            start_date = cls._parse_date(start_date_str)
            end_date = cls._parse_date(end_date_str)

            return cls._generate_date_range(start_date, end_date)

        date = cls._parse_date(input_date)
        return [date.strftime(cls._DATE_FORMAT)]


class ConfigLoader:
    _logger = CustomLogger.get_logger()

    def __init__(self, config_file_path: str = _CONFIG_FILE_PATH):
        self._config_file_path = config_file_path
        self._config: Dict[str, Any] = {}

    def _load_config(self) -> None:
        try:
            with open(self._config_file_path, "r") as file:
                self._config = yaml.safe_load(file)
            self._logger.info(
                f"Configuration file {self._config_file_path}"
                " loaded successfully."
            )
        except FileNotFoundError:
            raise FileNotFoundError(
                f"Configuration file {self._config_file_path} not found."
            )
        except yaml.YAMLError as e:
            raise yaml.YAMLError(
                f"Error parsing {self._config_file_path}: {e}."
            )

    def get(self, key: str, default: Any = None) -> Any:
        if not self._config:
            self._load_config()
        return self._config.get(key, default)

    def get_processing_time_threshold(self) -> int:
        threshold = self.get("processing_time_threshold_seconds")

        if isinstance(threshold, int):
            return threshold
        else:
            raise ValueError(
                "Invalid value for 'processing_time_threshold_seconds'"
                " in config file."
            )


class TargetHandler:
    @classmethod
    def get_target_prefixes(cls, config_loader: ConfigLoader) -> List[str]:
        TARGET_INDEX = 2
        if len(sys.argv) > TARGET_INDEX:
            targets = sys.argv[TARGET_INDEX].split(",")
        else:
            targets = config_loader.get("targets", [])
        return targets

    @classmethod
    def get_target_fullnames(cls, target_prefixes: List[str]) -> List[str]:
        target_fullnames = []
        target_folders = os.listdir(_TARGET_FOLDERS_BASE_PATH)

        for target_prefix in target_prefixes:
            matched_target_fullnames = [
                target_folder
                for target_folder in target_folders
                if target_folder.startswith(target_prefix)
            ]

            if matched_target_fullnames:
                target_fullnames.extend(matched_target_fullnames)
            else:
                raise ValueError(
                    f"No folder starting with target prefix '{target_prefix}'"
                    " was found in the log directory."
                )

        return target_fullnames


class CSVPathMapper:
    @staticmethod
    def get_targets_with_csv_path_for_each_date(
        date_range: List[str], target_fullnames: List[str]
    ) -> Dict[str, Dict[str, str | None]]:
        targets_and_csv_path_by_dates = {}

        for target_fullname in target_fullnames:
            csv_paths_by_date = {
                date: FileUtility.get_csv_path(
                    os.path.join(_TARGET_FOLDERS_BASE_PATH, target_fullname),
                    date,
                )
                for date in date_range
            }
            targets_and_csv_path_by_dates[target_fullname] = csv_paths_by_date

        return targets_and_csv_path_by_dates


class FileUtility:
    @staticmethod
    def create_excel_path(target_fullname: str) -> str:
        excel_name = f"{target_fullname}.xlsx"
        excel_path = os.path.join(_EXCEL_FOLDER_PATH, excel_name)
        return excel_path

    @staticmethod
    def create_directory(file_path: str) -> None:
        directory_for_file = os.path.dirname(file_path)
        os.makedirs(directory_for_file, exist_ok=True)

    @staticmethod
    def get_csv_path(target_folder_path: str, date: str) -> str | None:
        csv_name = f"test_{date}.csv"
        csv_path = os.path.join(target_folder_path, csv_name)

        if os.path.exists(csv_path):
            return csv_path
        else:
            return None


class CSVConsolidator:
    _logger = CustomLogger.get_logger()

    def __init__(self, writer: pd.ExcelWriter, workbook: Workbook) -> None:
        self._writer = writer
        self._workbook = workbook

        self._dates_with_merge_failure: set[str] = set()

    def _create_sentinel_sheet(self) -> None:
        pd.DataFrame({"A": ["SENTINEL_SHEET"]}).to_excel(
            self._writer,
            sheet_name="SENTINEL_SHEET",
            index=False,
            header=False,
        )

    def _create_sheet_from_csv(self, date: str, csv_path: str) -> None:
        try:
            df = pd.read_csv(csv_path)
            df.to_excel(self._writer, sheet_name=date, index=False)
        except Exception as e:
            self._logger.error(f"Failed to read CSV file at {csv_path}: {e}")
            self._dates_with_merge_failure.add(date)

    def _create_no_csv_sheet(self, date: str) -> None:
        df_for_no_csv = pd.DataFrame({"A": ["No CSV file found."]})
        df_for_no_csv.to_excel(
            self._writer, sheet_name=date, index=False, header=False
        )

        TRANSPARENT = "FF"
        GRAY = "7F7F7F"
        GRAY_WITH_TRANSPARENT = TRANSPARENT + GRAY
        self._writer.sheets[date].sheet_properties.tabColor = (
            GRAY_WITH_TRANSPARENT
        )

    def _create_sheets(
        self, csv_paths_for_each_date: dict[str, str | None]
    ) -> None:
        total_targets = len(csv_paths_for_each_date)

        for current_target_number, (date, csv_path) in enumerate(
            csv_paths_for_each_date.items(), start=1
        ):
            if csv_path:
                self._create_sheet_from_csv(date, csv_path)
            else:
                self._create_no_csv_sheet(date)

            self._logger.info(
                f"Added sheet: {date}."
                f" ({current_target_number}/{total_targets})"
            )

    def _delete_sentinel_sheet(self) -> None:
        if "SENTINEL_SHEET" in self._workbook.sheetnames:
            del self._workbook["SENTINEL_SHEET"]

    def consolidate_csvs_to_excel(
        self, csv_paths_for_each_date: dict[str, str | None]
    ) -> None:
        self._logger.info("Starting to merge.")

        self._create_sentinel_sheet()
        self._create_sheets(csv_paths_for_each_date)
        self._delete_sentinel_sheet()

        self._logger.info("Merging completed.")

    def get_dates_with_merge_failure(self) -> Dict[str, set[str]]:
        return {"dates_with_merge_failure": self._dates_with_merge_failure}


class ExcelAnalyzer:
    _logger = CustomLogger.get_logger()
    _TRANSPARENT = "FF"
    _YELLOW = "FFFF7F"
    _GRAY = "7F7F7F"
    _YELLOW_WITH_TRANSPARENT = _TRANSPARENT + _YELLOW
    _GRAY_WITH_TRANSPARENT = _TRANSPARENT + _GRAY

    def __init__(self, workbook: Workbook) -> None:
        self._workbook = workbook
        self._dates_with_threshold_exceedance: set[str] = set()
        self._dates_with_anomaly_value: set[str] = set()

    @staticmethod
    def _highlight_cell(cell: Cell, color_code: str) -> None:
        pattern_fill = PatternFill(start_color=color_code, fill_type="solid")
        cell.fill = pattern_fill

    @staticmethod
    def _calculate_color_based_on_excess_ratio(
        processing_time_seconds: int, threshold: int
    ) -> str:
        excess_ratio = (processing_time_seconds - threshold) / threshold
        clamped_excess_ratio = min(excess_ratio, 1)

        MAX_GREEN_VALUE = 255
        MIN_GREEN_VALUE = MAX_GREEN_VALUE / 2
        green_value = int(
            MAX_GREEN_VALUE
            - (MAX_GREEN_VALUE - MIN_GREEN_VALUE) * clamped_excess_ratio
        )

        green_hex_value = f"{green_value:02X}"
        color_code = f"FF{green_hex_value}7F"

        return color_code

    def _check_and_highlight_processing_time(
        self,
        processing_time_cell: Cell,
        threshold: int,
    ) -> bool:
        processing_time_value = processing_time_cell.value

        if processing_time_value:
            try:
                processing_time_seconds = int(
                    processing_time_value.rstrip("s")
                )

                if processing_time_seconds >= threshold:
                    color_code = self._calculate_color_based_on_excess_ratio(
                        processing_time_seconds, threshold
                    )
                    self._highlight_cell(processing_time_cell, color_code)
                    return True
            except ValueError:
                self._logger.warning(
                    f"Invalid processing time value: {processing_time_value}"
                )

        return False

    def _check_and_highlight_alert_detail(
        self,
        alert_detail_cell: Cell,
    ) -> bool:
        alert_detail_value = alert_detail_cell.value

        if alert_detail_value:
            try:
                alert_detail_data = json.loads(alert_detail_value)
                if any(
                    item.get("random_key") is True
                    for item in alert_detail_data
                ):
                    self._highlight_cell(
                        alert_detail_cell, self._YELLOW_WITH_TRANSPARENT
                    )
                    return True
            except json.JSONDecodeError:
                self._logger.warning(
                    f"Invalid JSON format found: {alert_detail_value}"
                )
        return False

    def _log_detected_anomalies(self, date: str) -> None:
        if date in self._dates_with_threshold_exceedance:
            self._logger.warning(
                "Exceeded processing time threshold detected"
                f" for date: {date}."
            )

        if date in self._dates_with_anomaly_value:
            self._logger.warning(f"Anomaly value detected for date: {date}.")

    def highlight_cells_and_sheet_tab_by_criteria(
        self, threshold: int
    ) -> None:
        self._logger.info("Starting to highlight.")
        total_sheets = len(self._workbook.sheetnames)

        for current_sheet_number, date in enumerate(
            self._workbook.sheetnames, start=1
        ):
            sheet = self._workbook[date]
            has_highlighted_cell = False

            for row in sheet.iter_rows(min_row=_DATA_START_ROW):
                processing_time_cell = row[_PROCESSING_TIME_COLUMN]
                alert_detail_cell = row[_ALERT_DETAIL_COLUMN]

                if self._check_and_highlight_processing_time(
                    processing_time_cell, threshold
                ):
                    self._dates_with_threshold_exceedance.add(date)
                    has_highlighted_cell = True

                if self._check_and_highlight_alert_detail(alert_detail_cell):
                    self._dates_with_anomaly_value.add(date)
                    has_highlighted_cell = True

            if has_highlighted_cell:
                sheet.sheet_properties.tabColor = self._YELLOW_WITH_TRANSPARENT
                self._log_detected_anomalies(date)

            self._logger.info(
                f"Analyzed sheet: {date}."
                f" ({current_sheet_number}/{total_sheets})"
            )
        self._logger.info("Highlighting completed.")

    def _create_new_order(self) -> List[str]:
        yellow_sheets = []
        gray_sheets = []
        other_sheets = []

        for date in self._workbook.sheetnames:
            sheet_tab_color = self._workbook[date].sheet_properties.tabColor

            if sheet_tab_color is None:
                other_sheets.append(date)
            else:
                sheet_color_value = sheet_tab_color.value

                if sheet_color_value == self._YELLOW_WITH_TRANSPARENT:
                    yellow_sheets.append(date)
                elif sheet_color_value == self._GRAY_WITH_TRANSPARENT:
                    gray_sheets.append(date)

        return yellow_sheets + other_sheets + gray_sheets

    def reorder_sheets_by_color(self) -> None:
        self._logger.info("Starting to reorder.")
        new_order = self._create_new_order()

        total_sheets = len(self._workbook.sheetnames)
        for current_sheet_number, date in enumerate(new_order, start=1):
            self._workbook.move_sheet(date, total_sheets)
            self._logger.info(
                f"Reordered sheet: {date}."
                f" ({current_sheet_number}/{total_sheets})"
            )
        self._logger.info("Reordering completed.")

    def get_analysis_results(self) -> Dict[str, set[str]]:
        return {
            "dates_with_threshold_exceedance": self._dates_with_threshold_exceedance,  # noqa E501
            "dates_with_anomaly_value": self._dates_with_anomaly_value,
        }


class ProcessingSummary:
    _logger = CustomLogger.get_logger()

    def __init__(self) -> None:
        self.daily_summaries: Dict[str, List[str]] = {}
        self.daily_processing_results: Dict[str, Dict[str, Set[str]]] = {}

    def add_missing_csv_info(
        self,
        targets_and_csv_path_by_dates: Dict[str, Dict[str, str | None]],
    ) -> None:
        for (
            target_fullname,
            csv_paths_by_date,
        ) in targets_and_csv_path_by_dates.items():
            for date, csv_path in csv_paths_by_date.items():
                if csv_path is None:
                    self.daily_summaries.setdefault(
                        target_fullname, []
                    ).append(f"Missing CSV for date {date}")

    def save_daily_processing_results(
        self,
        target: str,
        csv_consolidator: CSVConsolidator,
        excel_analyzer: ExcelAnalyzer,
    ) -> None:
        merge_failed_hosts = csv_consolidator.get_dates_with_merge_failure()
        analysis_results = excel_analyzer.get_analysis_results()

        self.daily_processing_results.setdefault(
            target,
            {
                "dates_with_merge_failure": set(),
                "dates_with_threshold_exceedance": set(),
                "dates_with_anomaly_value": set(),
            },
        )

        self.daily_processing_results[target][
            "dates_with_merge_failure"
        ].update(merge_failed_hosts["dates_with_merge_failure"])

        self.daily_processing_results[target][
            "dates_with_threshold_exceedance"
        ].update(analysis_results["dates_with_threshold_exceedance"])

        self.daily_processing_results[target][
            "dates_with_anomaly_value"
        ].update(analysis_results["dates_with_anomaly_value"])

    def _summarize_daily_processing_results(self) -> None:
        for date, summary in self.daily_processing_results.items():
            day_summary = []

            if summary.get("dates_with_threshold_exceedance"):
                dates_with_threshold_exceedance = summary[
                    "dates_with_threshold_exceedance"
                ]
                day_summary.append(
                    "Exceeded threshold detected for dates:"
                    f" {dates_with_threshold_exceedance}"
                )

            if summary.get("dates_with_anomaly_value"):
                dates_with_anomaly_value = summary["dates_with_anomaly_value"]
                day_summary.append(
                    "Anomaly value detected for dates:"
                    f" {dates_with_anomaly_value}"
                )

            if summary.get("dates_with_merge_failure"):
                merge_failed_hosts = summary["dates_with_merge_failure"]
                day_summary.append(f"Merge failed hosts: {merge_failed_hosts}")

            self.daily_summaries.setdefault(date, []).extend(day_summary)

    def log_daily_summaries(self) -> None:
        self._logger.info("Starting to log summary.")
        self._summarize_daily_processing_results()

        for target in sorted(self.daily_summaries.keys()):
            self._logger.info(f"Summary for {target}:")

            if not self.daily_summaries[target]:
                self._logger.info("No anomalies detected.")
            else:
                for summary_item in self.daily_summaries[target]:
                    self._logger.warning(f"{summary_item}")

        self._logger.info("Finished logging summary.")


def main() -> None:
    try:
        logger = CustomLogger.get_logger()
        logger.info("Process started.")

        date_range = DateHandler.get_date_range_or_yesterday()
        config_loader = ConfigLoader(_CONFIG_FILE_PATH)
        target_prefixes = TargetHandler.get_target_prefixes(config_loader)
        target_fullnames = TargetHandler.get_target_fullnames(target_prefixes)

        targets_with_csv_path_for_each_date = (
            CSVPathMapper.get_targets_with_csv_path_for_each_date(
                date_range, target_fullnames
            )
        )
        processing_time_threshold = (
            config_loader.get_processing_time_threshold()
        )
        processing_summary = ProcessingSummary()
        processing_summary.add_missing_csv_info(
            targets_with_csv_path_for_each_date
        )

        for target_fullname in target_fullnames:
            csv_paths_for_each_date = targets_with_csv_path_for_each_date.get(
                target_fullname
            )

            if csv_paths_for_each_date is None or all(
                csv_path is None
                for csv_path in csv_paths_for_each_date.values()
            ):
                logger.warning(
                    f"No CSV files found for host '{target_fullname}'."
                )
                continue

            excel_path = FileUtility.create_excel_path(target_fullname)
            FileUtility.create_directory(excel_path)

            logger.info(f"Starting to create '{excel_path}'.")
            with pd.ExcelWriter(
                excel_path, engine="openpyxl", mode="w"
            ) as writer:
                workbook = writer.book

                csv_consolidator = CSVConsolidator(writer, workbook)
                csv_consolidator.consolidate_csvs_to_excel(
                    csv_paths_for_each_date
                )

                excel_analyzer = ExcelAnalyzer(workbook)
                excel_analyzer.highlight_cells_and_sheet_tab_by_criteria(
                    processing_time_threshold
                )
                excel_analyzer.reorder_sheets_by_color()

                logger.info(f"Saving '{excel_path}'.")

            processing_summary.save_daily_processing_results(
                target_fullname,
                csv_consolidator,
                excel_analyzer,
            )
            logger.info(f"Finished creating '{excel_path}'.")

        processing_summary.log_daily_summaries()
        logger.info("Process completed.")
    except Exception as e:
        logger.error(f"An error occurred: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
