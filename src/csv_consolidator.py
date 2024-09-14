from abc import ABC, abstractmethod
from typing import Dict

import pandas as pd
from openpyxl import Workbook

from src.custom_logger import CustomLogger

_TRANSPARENT = "FF"
_GRAY = "7F7F7F"
_GRAY_WITH_TRANSPARENT = _TRANSPARENT + _GRAY


class _CSVConsolidator(ABC):
    _logger = CustomLogger.get_logger()

    def __init__(self, writer: pd.ExcelWriter, workbook: Workbook) -> None:
        self._writer = writer
        self._workbook = workbook

    def _create_sentinel_sheet(self) -> None:
        pd.DataFrame({"A": ["SENTINEL_SHEET"]}).to_excel(
            self._writer,
            sheet_name="SENTINEL_SHEET",
            index=False,
            header=False,
        )

    @abstractmethod
    def _create_sheet_from_csv(
        self, target_or_date: str, csv_path: str
    ) -> None:
        pass

    def _create_no_csv_sheet(self, date: str) -> None:
        df_for_no_csv = pd.DataFrame({"A": ["No CSV file found."]})
        df_for_no_csv.to_excel(
            self._writer, sheet_name=date, index=False, header=False
        )

        self._writer.sheets[date].sheet_properties.tabColor = (
            _GRAY_WITH_TRANSPARENT
        )

    def _create_sheets(
        self, csv_paths_for_each_date: dict[str, str | None]
    ) -> None:
        total_targets = len(csv_paths_for_each_date)

        for current_target_number, (date, csv_path) in enumerate(
            csv_paths_for_each_date.items(), start=1
        ):
            if csv_path:
                self._create_sheet_from_csv(date, csv_path)
            else:
                self._create_no_csv_sheet(date)

            self._logger.info(
                f"Added sheet: {date}."
                f" ({current_target_number}/{total_targets})"
            )

    def _delete_sentinel_sheet(self) -> None:
        if "SENTINEL_SHEET" in self._workbook.sheetnames:
            del self._workbook["SENTINEL_SHEET"]

    def consolidate_csvs_to_excel(
        self, csv_paths_for_each_date: dict[str, str | None]
    ) -> None:
        self._logger.info("Starting to merge.")

        self._create_sentinel_sheet()
        self._create_sheets(csv_paths_for_each_date)
        self._delete_sentinel_sheet()

        self._logger.info("Merging completed.")


class DateBasedCSVConsolidator(_CSVConsolidator):
    def __init__(self, writer: pd.ExcelWriter, workbook: Workbook) -> None:
        super().__init__(writer, workbook)
        self._merge_failed_hosts: set[str] = set()

    def _create_sheet_from_csv(self, target_name: str, csv_path: str) -> None:
        try:
            df = pd.read_csv(csv_path)
            df.to_excel(self._writer, sheet_name=target_name, index=False)
        except Exception as e:
            self._logger.error(f"Failed to read CSV file at {csv_path}: {e}")
            self._merge_failed_hosts.add(target_name)

    def get_merge_failed_hosts(self) -> Dict[str, set[str]]:
        return {"merge_failed_hosts": self._merge_failed_hosts}


class HostBasedCSVConsolidator(_CSVConsolidator):
    def __init__(self, writer: pd.ExcelWriter, workbook: Workbook) -> None:
        super().__init__(writer, workbook)
        self._dates_with_merge_failure: set[str] = set()

    def _create_sheet_from_csv(self, date: str, csv_path: str) -> None:
        try:
            df = pd.read_csv(csv_path)
            df.to_excel(self._writer, sheet_name=date, index=False)
        except Exception as e:
            self._logger.error(f"Failed to read CSV file at {csv_path}: {e}")
            self._dates_with_merge_failure.add(date)

    def get_dates_with_merge_failure(self) -> Dict[str, set[str]]:
        return {"dates_with_merge_failure": self._dates_with_merge_failure}
