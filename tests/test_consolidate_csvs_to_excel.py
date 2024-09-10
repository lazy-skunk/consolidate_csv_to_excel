import json
import logging
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Type
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

from src.consolidate_csvs_to_excel import (
    ConfigLoader,
    CSVConsolidator,
    CSVPathMapper,
    DateHandler,
    ExcelAnalyzer,
    FileUtility,
    TargetHandler,
)

_DATE_FORMAT = "%Y%m%d"
_YESTERDAY = (datetime.now() - timedelta(days=1)).strftime(_DATE_FORMAT)
_TOMORROW = (datetime.now() + timedelta(days=1)).strftime(_DATE_FORMAT)
_GRAY_WITH_TRANSPARENT = "007F7F7F"


@pytest.mark.parametrize(
    "argv, expected",
    [
        (
            ["test.py"],
            [_YESTERDAY],
        ),
        (["test.py", "19880209"], ["19880209"]),
        (["test.py", "19880209~19880209"], ["19880209"]),
        (
            ["test.py", "19880209~19880211"],
            ["19880209", "19880210", "19880211"],
        ),
        (
            ["test.py", "19880211~19880209"],
            ["19880209", "19880210", "19880211"],
        ),
    ],
)
def test_get_date_range_or_yesterday(
    argv: List[str],
    expected: List[str],
) -> None:
    with patch.object(sys, "argv", argv):
        result = DateHandler.get_date_range_or_yesterday()
        assert result == expected


@pytest.mark.parametrize(
    "argv",
    [
        (["test.py", "1988029"]),
        (["test.py", "1988-02-09"]),
        (["test.py", "1988~02~09"]),
        (["test.py", _TOMORROW]),
        (["test.py", "invalid_date"]),
    ],
)
def test_get_date_range_or_yesterday_with_invalid_dates(
    argv: List[str],
) -> None:
    with patch.object(sys, "argv", argv):
        with pytest.raises(ValueError):
            DateHandler.get_date_range_or_yesterday()


def test_get_processing_time_threshold() -> None:
    temp_config_path = os.path.join("tests", "data", "test_config.yml")
    config_loader = ConfigLoader(temp_config_path)
    mock_logger = MagicMock(spec=logging.Logger)

    with patch.object(config_loader, "_logger", mock_logger):
        threshold = config_loader.get_processing_time_threshold()
        expected = 4
        assert threshold == expected


def test_get_processing_time_threshold_with_nonexistent_file() -> None:
    mock_logger = MagicMock(spec=logging.Logger)
    config_loader = ConfigLoader("NONEXISTENT_CONFIG.YAML")

    with patch.object(config_loader, "_logger", mock_logger):
        with pytest.raises(FileNotFoundError):
            config_loader.get_processing_time_threshold()


def test_get_processing_time_threshold_with_invalid_config() -> None:
    invalid_config_path = os.path.join("tests", "data", "invalid_yaml.yml")
    config_loader = ConfigLoader(invalid_config_path)
    mock_logger = MagicMock(spec=logging.Logger)

    with patch.object(config_loader, "_logger", mock_logger):
        with pytest.raises(yaml.YAMLError):
            config_loader.get_processing_time_threshold()


def test_get_processing_time_threshold_with_invalid_threshold() -> None:
    invalid_threshold_path = os.path.join(
        "tests", "data", "invalid_threshold.yml"
    )
    config_loader = ConfigLoader(invalid_threshold_path)
    mock_logger = MagicMock(spec=logging.Logger)

    with patch.object(config_loader, "_logger", mock_logger):
        with pytest.raises(ValueError):
            config_loader.get_processing_time_threshold()


@pytest.mark.parametrize(
    "argv, config_targets, expected",
    [
        (
            ["test.py", "19880209", "target1,target2"],
            None,
            ["target1", "target2"],
        ),
        (
            ["test.py"],
            ["config_target1", "config_target2"],
            ["config_target1", "config_target2"],
        ),
    ],
)
def test_get_target_prefixes(
    argv: List[str],
    config_targets: List[str] | None,
    expected: List[str],
) -> None:
    mock_config_loader = MagicMock(spec=ConfigLoader)
    if config_targets:
        mock_config_loader.get.return_value = config_targets

    with patch("sys.argv", argv):
        target_prefixes = TargetHandler.get_target_prefixes(mock_config_loader)
        assert target_prefixes == expected


@pytest.mark.parametrize(
    "target_prefixes, expected",
    [
        (["target"], ["target_0", "target_1", "target_2", "target_3"]),
        (["target_2"], ["target_2"]),
    ],
)
def test_get_target_fullnames(
    target_prefixes: List[str],
    expected: List[str],
) -> None:
    test_folders_base_path = os.path.join("tests", "data")
    with patch(
        "src.consolidate_csvs_to_excel._TARGET_FOLDERS_BASE_PATH",
        test_folders_base_path,
    ):
        host_fullnames = TargetHandler.get_target_fullnames(target_prefixes)
        assert host_fullnames == expected


def test_get_target_fullnames_with_nonexistent_target() -> None:
    test_folders_base_path = os.path.join("tests", "data")
    with patch(
        "src.consolidate_csvs_to_excel._TARGET_FOLDERS_BASE_PATH",
        test_folders_base_path,
    ):
        with pytest.raises(ValueError):
            TargetHandler.get_target_fullnames(["NONEXISTENT_TARGET"])


@pytest.mark.parametrize(
    "date_range, target_fullnames",
    [
        (["19880209", "19880210"], ["target_0", "target_1"]),
        (["19880209"], ["target_2"]),
    ],
)
def test_get_targets_and_csv_path_by_dates(
    date_range: List[str],
    target_fullnames: List[str],
) -> None:
    test_folders_base_path = os.path.join("tests", "data")

    expected: Dict[str, Dict[str, str]] = {}
    for date in date_range:
        expected[date] = {}
        for target_fullname in target_fullnames:
            expected[date][target_fullname] = os.path.join(
                test_folders_base_path, target_fullname, f"test_{date}.csv"
            )

    with patch(
        "src.consolidate_csvs_to_excel._TARGET_FOLDERS_BASE_PATH",
        test_folders_base_path,
    ):
        result = CSVPathMapper.get_targets_and_csv_path_by_dates(
            date_range, target_fullnames
        )

    assert result == expected


# @pytest.mark.parametrize(
#     "argv, targets, expected",
#     [
#         (["test.py"], ["target1", "target2"], "config"),
#         (
#             ["test.py", "19880209"],
#             ["target1", "target2"],
#             "config",
#         ),
#         (
#             ["test.py", "19880209", "target"],
#             ["target1", "target2"],
#             "target1_target2",
#         ),
#     ],
# )
# def test_create_file_name_suffix(
#     argv: List[str], targets: List[str], expected: str
# ) -> None:
#     with patch.object(sys, "argv", argv):
#         result = FileUtility.create_file_name_suffix(targets)
#         assert result == expected


# @pytest.mark.parametrize(
#     "argv, suffix",
#     [
#         (["test.py", "19880209", "target"], "target1_target2"),
#         (["test.py", "19880209"], "config"),
#         (["test.py"], "config"),
#     ],
# )
# def test_create_excel_path(
#     tmp_path: Path, argv: List[str], suffix: str
# ) -> None:
#     DATE = "19880209"

#     with (
#         patch("sys.argv", argv),
#         patch(
#             "src.consolidate_csv_to_excel._EXCEL_FOLDER_PATH",
#             tmp_path,
#         ),
#     ):
#         expected = os.path.join(
#             tmp_path,
#             DATE,
#             f"{DATE}_{suffix}.xlsx",
#         )

#         result = FileUtility.create_excel_path(DATE, suffix)
#         assert result == expected


# def test_create_excel_directory(tmp_path: str) -> None:
#     excel_file_path = os.path.join(tmp_path, "19880209", "19880209_file.xlsx")
#     excel_directory = os.path.dirname(excel_file_path)

#     assert not os.path.exists(excel_directory)
#     FileUtility.create_directory(excel_file_path)
#     assert os.path.exists(excel_directory)


# @pytest.mark.parametrize(
#     "target_folder_path, date, file_exists, expected",
#     [
#         (
#             "/path/to/target",
#             "19880209",
#             True,
#             "/path/to/target/test_19880209.csv",
#         ),
#         (
#             "/path/to/target",
#             "19880209",
#             False,
#             None,
#         ),
#     ],
# )
# def test_get_merged_csv_path(
#     target_folder_path: str,
#     date: str,
#     file_exists: bool,
#     expected: str | None,
# ) -> None:
#     with patch("os.path.exists", return_value=file_exists):
#         result = FileUtility.get_merged_csv_path(target_folder_path, date)
#         assert result == expected


# def test_create_sentinel_sheet(tmp_path: Path) -> None:
#     mock_logger = MagicMock(spec=logging.Logger)
#     tmp_excel = os.path.join(tmp_path, "excel.xlsx")
#     with pd.ExcelWriter(tmp_excel, engine="openpyxl", mode="w") as writer:
#         workbook = writer.book
#         csv_consolidator = CSVConsolidator(writer, workbook, mock_logger)
#         csv_consolidator._create_sentinel_sheet()

#         assert "SENTINEL_SHEET" in workbook.sheetnames


# @pytest.mark.parametrize(
#     "date, no_csv_found, exception",
#     [
#         ("19880209", False, None),
#         ("INVALID_DATE", True, None),
#         ("19880209", False, Exception),
#     ],
# )
# def test_search_and_append_csv_to_excel(
#     tmp_path: Path,
#     date: str,
#     no_csv_found: bool,
#     exception: type[Exception] | None,
# ) -> None:
#     TestHelper.prepare_tmp_four_csvs(tmp_path)

#     mock_logger = MagicMock(spec=logging.Logger)
#     tmp_excel = os.path.join(tmp_path, "excel.xlsx")

#     with (
#         patch(
#             "src.consolidate_csv_to_excel._TARGET_FOLDERS_BASE_PATH",
#             f"{tmp_path}",
#         ),
#         pd.ExcelWriter(tmp_excel, engine="openpyxl", mode="w") as writer,
#     ):
#         workbook = writer.book
#         csv_consolidator = CSVConsolidator(writer, workbook, mock_logger)

#         pd.DataFrame({"A": ["SENTINEL_SHEET"]}).to_excel(
#             writer, sheet_name="SENTINEL_SHEET", index=False, header=False
#         )

#         target_fullnames = [f"target_{i}" for i in range(4)]

#         if exception:
#             with patch("pandas.read_csv", side_effect=exception):
#                 csv_consolidator._create_sheet(date, target_fullnames)
#                 assert csv_consolidator._copied_count == 0
#                 assert csv_consolidator._no_csv_count == 0
#                 assert csv_consolidator._failed_count == 4
#                 assert len(csv_consolidator._merge_failed_hosts) == 4
#                 return
#         else:
#             csv_consolidator._create_sheet(date, target_fullnames)

#         assert set(workbook.sheetnames) == {
#             "SENTINEL_SHEET",
#             "target_0",
#             "target_1",
#             "target_2",
#             "target_3",
#         }

#         if no_csv_found:
#             for sheet_name in target_fullnames:
#                 sheet = workbook[sheet_name]
#                 assert (
#                     sheet.sheet_properties.tabColor.value
#                     == TestHelper.GRAY_WITH_TRANSPARENT
#                 )
#             assert csv_consolidator._copied_count == 0
#             assert csv_consolidator._no_csv_count == 4
#             assert csv_consolidator._failed_count == 0
#             assert len(csv_consolidator._merge_failed_hosts) == 0
#         else:
#             for sheet_name in target_fullnames:
#                 sheet = workbook[sheet_name]
#                 assert sheet.sheet_properties.tabColor is None
#             assert csv_consolidator._copied_count == 4
#             assert csv_consolidator._no_csv_count == 0
#             assert csv_consolidator._failed_count == 0
#             assert len(csv_consolidator._merge_failed_hosts) == 0


# def test_remove_sentinel_sheet_exists(tmp_path: Path) -> None:
#     mock_logger = MagicMock(spec=logging.Logger)
#     tmp_excel = os.path.join(tmp_path, "excel.xlsx")

#     with pd.ExcelWriter(tmp_excel, engine="openpyxl", mode="w") as writer:
#         workbook = writer.book
#         csv_consolidator = CSVConsolidator(writer, workbook, mock_logger)

#         pd.DataFrame({"A": ["SENTINEL_SHEET"]}).to_excel(
#             writer, sheet_name="SENTINEL_SHEET", index=False, header=False
#         )
#         pd.DataFrame({"A": ["OTHER_SHEET"]}).to_excel(
#             writer, sheet_name="OTHER_SHEET", index=False, header=False
#         )

#         csv_consolidator._delete_sentinel_sheet()
#         assert "SENTINEL_SHEET" not in workbook.sheetnames
#         assert "OTHER_SHEET" in workbook.sheetnames


# def test_get_summary(csv_consolidator: CSVConsolidator) -> None:
#     summary = csv_consolidator.get_summary()
#     assert summary == {
#         "copied": 0,
#         "no_csv": 0,
#         "failed": 0,
#         "failed_hosts": [],
#     }

#     csv_consolidator._copied_count = 1
#     csv_consolidator._no_csv_count = 2
#     csv_consolidator._failed_count = 3
#     csv_consolidator._failed_hosts = ["host1", "host2"]

#     summary = csv_consolidator.get_summary()
#     assert summary == {
#         "copied": 1,
#         "no_csv": 2,
#         "failed": 3,
#         "failed_hosts": ["host1", "host2"],
#     }


# def test_highlight_cells_and_sheet_tabs(
#     excel_analyzer: ExcelAnalyzer,
#     prepare_tmp_excel: None,
#     tmp_path_for_excel: str,
# ) -> None:
#     threshold = 4  # 0, 2, 4, 6

#     excel_analyzer.highlight_cells_and_sheet_tab_by_criteria(
#         tmp_path_for_excel, threshold
#     )
#     NO_COLOR_WITH_TRANSPARENT = "00000000"
#     workbook = load_workbook(tmp_path_for_excel)
#     try:
#         for sheet_name in workbook.sheetnames:
#             sheet = workbook[sheet_name]

#             has_yellow_tab = sheet.sheet_properties.tabColor is not None

#             has_highlighted_cells = any(
#                 cell.fill.start_color.rgb != NO_COLOR_WITH_TRANSPARENT
#                 for row in sheet.iter_rows(min_row=2)
#                 for cell in row
#             )

#             if has_yellow_tab:
#                 assert has_highlighted_cells
#             else:
#                 assert not has_highlighted_cells
#     finally:
#         workbook.close()


# def test_reorder_sheets_by_color(
#     excel_analyzer: ExcelAnalyzer,
#     prepare_tmp_excel_for_reordering: None,
#     tmp_path_for_excel: str,
# ) -> None:
#     try:
#         workbook = load_workbook(tmp_path_for_excel)
#         initial_order = workbook.sheetnames
#         assert initial_order == ["Other_Sheet", "Gray_Sheet", "Yellow_Sheet"]

#         excel_analyzer.reorder_sheets_by_color(tmp_path_for_excel)

#         workbook = load_workbook(tmp_path_for_excel)
#         reordered_sheet_names = workbook.sheetnames

#         expected_order = ["Yellow_Sheet", "Other_Sheet", "Gray_Sheet"]

#         assert reordered_sheet_names == expected_order
#     finally:
#         workbook.close()


# def test_get_hosts_to_check(
#     excel_analyzer: ExcelAnalyzer,
# ) -> None:
#     excel_analyzer._hosts_to_check = {"target_1", "target_2", "target_3"}

#     hosts_to_check = excel_analyzer.get_hosts_to_check()

#     assert hosts_to_check == {"target_1", "target_2", "target_3"}
