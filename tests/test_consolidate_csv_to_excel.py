import logging
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Type
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

from src.consolidate_csv_to_excel import (
    ConfigLoader,
    CSVConsolidator,
    DateHandler,
    ExcelAnalyzer,
    TargetHandler,
)


class TestHelper:
    DATE_FORMAT = "%Y%m%d"
    YESTERDAY = (datetime.now() - timedelta(days=1)).strftime(DATE_FORMAT)
    TOMORROW = (datetime.now() + timedelta(days=1)).strftime(DATE_FORMAT)
    GRAY_WITH_TRANSPARENT = "007F7F7F"

    @staticmethod
    def create_temp_config_and_return_path(
        tmp_path: Path, config_data: dict
    ) -> str:
        temp_config_path = os.path.join(tmp_path, "temp_config.yaml")

        with open(temp_config_path, "w") as file:
            yaml.dump(config_data, file)

        return temp_config_path

    @staticmethod
    def create_malformed_config_and_return_path(tmp_path: Path) -> str:
        malformed_content = "invalid_yaml: [unclosed list"
        temp_file = os.path.join(tmp_path, "malformed_config.yaml")

        with open(temp_file, "w") as file:
            file.write(malformed_content)

        return temp_file


@pytest.mark.parametrize(
    "argv, expected",
    [
        (
            ["test.py"],
            [TestHelper.YESTERDAY],
        ),
        (["test.py", "19880209"], ["19880209"]),
        (["test.py", "19880209~19880209"], ["19880209"]),
        (
            ["test.py", "19880209~19880211"],
            ["19880209", "19880210", "19880211"],
        ),
        (
            ["test.py", "19880211~19880209"],
            ["19880209", "19880210", "19880211"],
        ),
    ],
)
def test_get_date_range_or_yesterday(
    argv: List[str],
    expected: List[str],
) -> None:
    mock_logger = MagicMock(spec=logging.Logger)
    date_handler = DateHandler(mock_logger)

    with patch.object(sys, "argv", argv):
        result = date_handler.get_date_range_or_yesterday()
        assert result == expected


@pytest.mark.parametrize(
    "argv",
    [
        (["test.py", "1988029"]),
        (["test.py", "1988-02-09"]),
        (["test.py", "1988~02~09"]),
        (
            [
                "test.py",
                TestHelper.TOMORROW,
            ]
        ),
        (["test.py", "invalid_date"]),
    ],
)
def test_get_date_range_or_yesterday_with_invalid_dates(
    argv: List[str],
) -> None:
    mock_logger = MagicMock(spec=logging.Logger)
    date_handler = DateHandler(mock_logger)

    with patch.object(sys, "argv", argv):
        with pytest.raises(ValueError):
            date_handler.get_date_range_or_yesterday()


@pytest.mark.parametrize(
    "config_data, expected, exception",
    [
        ({"processing_time_threshold_seconds": 30}, 30, None),
        (
            {"processing_time_threshold_seconds": "invalid"},
            None,
            ValueError,
        ),
        ({}, None, ValueError),
    ],
)
def test_get_processing_time_threshold(
    tmp_path: Path,
    config_data: dict,
    expected: int,
    exception: type[ValueError] | None,
) -> None:
    mock_logger = MagicMock(spec=logging.Logger)
    temp_config_path = TestHelper.create_temp_config_and_return_path(
        tmp_path, config_data
    )
    config_loader = ConfigLoader(mock_logger, temp_config_path)

    if exception:
        with pytest.raises(exception):
            config_loader.get_processing_time_threshold()
    else:
        threshold = config_loader.get_processing_time_threshold()
        assert threshold == expected


def test_config_not_found() -> None:
    mock_logger = MagicMock(spec=logging.Logger)
    with pytest.raises(FileNotFoundError):
        ConfigLoader(mock_logger, "non_existent_file.yaml")


def test_get_processing_time_threshold_with_malformed_config(
    tmp_path: Path,
) -> None:
    mock_logger = MagicMock(spec=logging.Logger)
    malformed_config_path = TestHelper.create_malformed_config_and_return_path(
        tmp_path
    )
    with pytest.raises(yaml.YAMLError):
        ConfigLoader(mock_logger, malformed_config_path)


@pytest.mark.parametrize(
    "argv, config_targets, expected",
    [
        (
            ["test.py", "dummy_arg", "target1,target2"],
            None,
            ["target1", "target2"],
        ),
        (
            ["test.py"],
            ["config_target1", "config_target2"],
            ["config_target1", "config_target2"],
        ),
    ],
)
def test_get_targets(
    argv: List[str],
    config_targets: List[str] | None,
    expected: List[str],
) -> None:
    mock_config_loader = MagicMock(spec=ConfigLoader)
    mock_logger = MagicMock(spec=logging.Logger)
    target_handler = TargetHandler(mock_config_loader, mock_logger)

    with patch("sys.argv", argv):
        if config_targets:
            mock_config_loader.get.return_value = config_targets

        targets = target_handler.get_targets()
        assert targets == expected


@pytest.mark.parametrize(
    "host_folders, targets, expected, exception",
    [
        (
            ["host1_log", "host2_log", "host3_log"],
            ["host1", "host2"],
            ["host1_log", "host2_log"],
            None,
        ),
        (
            ["host1_log", "host2_log", "host3_log"],
            ["host4"],
            [],
            ValueError,
        ),
        (
            ["host1_log", "host2_log", "host3_log"],
            ["host1", "host4"],
            ["host1_log"],
            None,
        ),
    ],
)
def test_get_existing_host_fullnames(
    host_folders: List[str],
    targets: List[str],
    expected: List[str],
    exception: Type[ValueError] | None,
) -> None:
    mock_config_loader = MagicMock(spec=ConfigLoader)
    mock_logger = MagicMock(spec=logging.Logger)
    target_handler = TargetHandler(mock_config_loader, mock_logger)

    with patch("os.listdir", return_value=host_folders):
        if exception:
            with pytest.raises(exception):
                target_handler.get_existing_target_fullnames(targets)
        else:
            host_fullnames = target_handler.get_existing_target_fullnames(
                targets
            )
            assert host_fullnames == expected


# @pytest.mark.parametrize(
#     "argv, expected_suffix",
#     [
#         (["test.py", "arg1", "arg2"], "target1_target2"),
#         (["test.py"], "config"),
#     ],
# )
# def test_create_excel_file_path(
#     csv_consolidator: CSVConsolidator,
#     tmp_path: Path,
#     argv: List[str],
#     expected_suffix: str,
# ) -> None:
#     date = "19880209"
#     targets = ["target1", "target2"]

#     with (
#         patch("sys.argv", argv),
#         patch(
#             "src.consolidate_csv_to_excel._EXCEL_FOLDER_PATH",
#             tmp_path,
#         ),
#     ):
#         expected = os.path.join(
#             tmp_path,
#             date,
#             f"{date}_{expected_suffix}.xlsx",
#         )

#         result = csv_consolidator.create_excel_file_path(date, targets)
#         assert result == expected


# def test_create_excel_directory(
#     csv_consolidator: CSVConsolidator, tmp_path: str
# ) -> None:
#     excel_file_path = os.path.join(tmp_path, "test", "file.xlsx")
#     excel_directory = os.path.dirname(excel_file_path)

#     assert not os.path.exists(excel_directory)
#     csv_consolidator.create_excel_directory(excel_file_path)
#     assert os.path.exists(excel_directory)


# def test_create_sentinel_sheet(
#     csv_consolidator: CSVConsolidator, tmp_path_for_excel: str
# ) -> None:
#     with pd.ExcelWriter(
#         tmp_path_for_excel, engine="openpyxl", mode="w"
#     ) as writer:
#         csv_consolidator.create_sentinel_sheet(writer)


# @pytest.mark.parametrize(
#     "date, no_csv_found, exception",
#     [
#         ("19880209", False, None),
#         ("INVALID_DATE", True, None),
#         ("19880209", False, Exception),
#     ],
# )
# def test_search_and_append_csv_to_excel(
#     csv_consolidator: CSVConsolidator,
#     prepare_tmp_csv: None,
#     prepare_tmp_excel_with_sentinel: None,
#     tmp_path: Path,
#     tmp_path_for_excel: str,
#     date: str,
#     no_csv_found: bool,
#     exception: type[Exception] | None,
# ) -> None:
#     target_fullnames = [f"target_{i}" for i in range(4)]

#     with patch(
#         "src.consolidate_csv_to_excel._TARGET_FOLDERS_BASE_PATH", f"{tmp_path}"
#     ):
#         if exception:
#             with patch("pandas.read_csv", side_effect=exception):
#                 csv_consolidator.search_and_append_csv_to_excel(
#                     date, target_fullnames, tmp_path_for_excel
#                 )
#                 assert csv_consolidator._copied_count == 0
#                 assert csv_consolidator._no_csv_count == 0
#                 assert csv_consolidator._failed_count == 4
#                 assert len(csv_consolidator._failed_hosts) == 4
#                 return
#         else:
#             csv_consolidator.search_and_append_csv_to_excel(
#                 date, target_fullnames, tmp_path_for_excel
#             )

#     try:
#         workbook = load_workbook(tmp_path_for_excel)

#         assert set(workbook.sheetnames) == {
#             "SENTINEL_SHEET",
#             "target_0",
#             "target_1",
#             "target_2",
#             "target_3",
#         }

#         if no_csv_found:
#             for sheet_name in target_fullnames:
#                 sheet = workbook[sheet_name]
#                 assert (
#                     sheet.sheet_properties.tabColor.value
#                     == TestHelper.GRAY_WITH_TRANSPARENT
#                 )
#             assert csv_consolidator._copied_count == 0
#             assert csv_consolidator._no_csv_count == 4
#             assert csv_consolidator._failed_count == 0
#             assert len(csv_consolidator._failed_hosts) == 0
#         else:
#             for sheet_name in target_fullnames:
#                 sheet = workbook[sheet_name]
#                 assert sheet.sheet_properties.tabColor is None
#             assert csv_consolidator._copied_count == 4
#             assert csv_consolidator._no_csv_count == 0
#             assert csv_consolidator._failed_count == 0
#             assert len(csv_consolidator._failed_hosts) == 0
#     finally:
#         workbook.close()


# def test_remove_sentinel_sheet_exists(
#     csv_consolidator: CSVConsolidator,
#     mock_logger: MagicMock,
#     prepare_tmp_excel_with_sentinel_and_dummy: None,
#     tmp_path_for_excel: str,
# ) -> None:
#     csv_consolidator.delete_sentinel_sheet(tmp_path_for_excel)

#     try:
#         workbook = load_workbook(tmp_path_for_excel)
#         assert "SENTINEL_SHEET" not in workbook.sheetnames
#     finally:
#         workbook.close()

#     csv_consolidator.delete_sentinel_sheet(tmp_path_for_excel)
#     mock_logger.warning.assert_called_once_with(
#         f"SENTINEL_SHEET not found in {tmp_path_for_excel}."
#     )


# def test_get_summary(csv_consolidator: CSVConsolidator) -> None:
#     summary = csv_consolidator.get_summary()
#     assert summary == {
#         "copied": 0,
#         "no_csv": 0,
#         "failed": 0,
#         "failed_hosts": [],
#     }

#     csv_consolidator._copied_count = 1
#     csv_consolidator._no_csv_count = 2
#     csv_consolidator._failed_count = 3
#     csv_consolidator._failed_hosts = ["host1", "host2"]

#     summary = csv_consolidator.get_summary()
#     assert summary == {
#         "copied": 1,
#         "no_csv": 2,
#         "failed": 3,
#         "failed_hosts": ["host1", "host2"],
#     }


# def test_highlight_cells_and_sheet_tabs(
#     excel_analyzer: ExcelAnalyzer,
#     prepare_tmp_excel: None,
#     tmp_path_for_excel: str,
# ) -> None:
#     threshold = 4  # 0, 2, 4, 6

#     excel_analyzer.highlight_cells_and_sheet_tab_by_criteria(
#         tmp_path_for_excel, threshold
#     )
#     NO_COLOR_WITH_TRANSPARENT = "00000000"
#     workbook = load_workbook(tmp_path_for_excel)
#     try:
#         for sheet_name in workbook.sheetnames:
#             sheet = workbook[sheet_name]

#             has_yellow_tab = sheet.sheet_properties.tabColor is not None

#             has_highlighted_cells = any(
#                 cell.fill.start_color.rgb != NO_COLOR_WITH_TRANSPARENT
#                 for row in sheet.iter_rows(min_row=2)
#                 for cell in row
#             )

#             if has_yellow_tab:
#                 assert has_highlighted_cells
#             else:
#                 assert not has_highlighted_cells
#     finally:
#         workbook.close()


# def test_reorder_sheets_by_color(
#     excel_analyzer: ExcelAnalyzer,
#     prepare_tmp_excel_for_reordering: None,
#     tmp_path_for_excel: str,
# ) -> None:
#     try:
#         workbook = load_workbook(tmp_path_for_excel)
#         initial_order = workbook.sheetnames
#         assert initial_order == ["Other_Sheet", "Gray_Sheet", "Yellow_Sheet"]

#         excel_analyzer.reorder_sheets_by_color(tmp_path_for_excel)

#         workbook = load_workbook(tmp_path_for_excel)
#         reordered_sheet_names = workbook.sheetnames

#         expected_order = ["Yellow_Sheet", "Other_Sheet", "Gray_Sheet"]

#         assert reordered_sheet_names == expected_order
#     finally:
#         workbook.close()


# def test_get_hosts_to_check(
#     excel_analyzer: ExcelAnalyzer,
# ) -> None:
#     excel_analyzer._hosts_to_check = {"target_1", "target_2", "target_3"}

#     hosts_to_check = excel_analyzer.get_hosts_to_check()

#     assert hosts_to_check == {"target_1", "target_2", "target_3"}
